from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.analytics import build_summary


NAVY = "0F172A"
BLUE = "2563EB"
WHITE = "FFFFFF"


def export_data(
    repositories: list[dict[str, Any]],
    output_dir: Path,
    username: str,
    export_format: str,
    logger: logging.Logger,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_username = username.replace("/", "_")
    dataframe = pd.DataFrame(repositories)
    summary = build_summary(repositories)

    if export_format in {"all", "json"}:
        _export_json(
            output_dir / f"{safe_username}_repositories.json",
            repositories,
            summary,
        )
        logger.info("JSON gerado.")

    if export_format in {"all", "csv"}:
        _export_csv(
            output_dir / f"{safe_username}_repositories.csv",
            dataframe,
        )
        logger.info("CSV gerado.")

    if export_format in {"all", "xlsx"}:
        _export_excel(
            output_dir / f"{safe_username}_repositories.xlsx",
            dataframe,
            summary,
        )
        logger.info("Excel gerado.")


def _export_json(
    filepath: Path,
    repositories: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    payload = {
        "summary": summary,
        "repositories": repositories,
    }

    filepath.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _export_csv(filepath: Path, dataframe: pd.DataFrame) -> None:
    dataframe.to_csv(
        filepath,
        index=False,
        encoding="utf-8-sig",
    )


def _export_excel(
    filepath: Path,
    dataframe: pd.DataFrame,
    summary: dict[str, Any],
) -> None:
    summary_rows = [
        {"Indicador": "Total de Repositórios", "Valor": summary["total_repositories"]},
        {"Indicador": "Total de Stars", "Valor": summary["total_stars"]},
        {"Indicador": "Total de Forks", "Valor": summary["total_forks"]},
        {"Indicador": "Issues Abertas", "Valor": summary["total_open_issues"]},
        {"Indicador": "Repositórios Arquivados", "Valor": summary["archived_repositories"]},
        {"Indicador": "Linguagem Mais Usada", "Valor": summary["most_used_language"]},
    ]

    language_rows = [
        {"Linguagem": language, "Repositórios": count}
        for language, count in summary["languages"].items()
    ]

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        dataframe.to_excel(writer, sheet_name="Repositorios", index=False)
        pd.DataFrame(summary_rows).to_excel(
            writer,
            sheet_name="Resumo",
            index=False,
        )
        pd.DataFrame(language_rows).to_excel(
            writer,
            sheet_name="Linguagens",
            index=False,
        )

    workbook = load_workbook(filepath)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column_cells in sheet.columns:
            max_length = max(
                len(str(cell.value or ""))
                for cell in column_cells
            )
            column_letter = get_column_letter(column_cells[0].column)
            sheet.column_dimensions[column_letter].width = min(
                max(max_length + 2, 12),
                42,
            )

    workbook.save(filepath)
