from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY = "0F172A"
WHITE = "FFFFFF"


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24


def _autofit(ws, max_width: int = 28) -> None:
    for cells in ws.columns:
        letter = get_column_letter(cells[0].column)
        longest = max(len(str(c.value or "")) for c in cells)
        ws.column_dimensions[letter].width = min(max(longest + 2, 11), max_width)


def _currency_column(ws, col: int, start_row: int = 2) -> None:
    for row in range(start_row, ws.max_row + 1):
        ws.cell(row=row, column=col).number_format = 'R$ #,##0.00'


def export_report(output_path: Path, treated, summary, sellers, products, regions, months) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        treated.to_excel(writer, sheet_name="Dados Tratados", index=False)
        pd.DataFrame([{"Indicador": k, "Valor": v} for k, v in summary.items()]).to_excel(
            writer, sheet_name="Resumo", index=False
        )
        sellers.to_excel(writer, sheet_name="Vendedores", index=False)
        products.to_excel(writer, sheet_name="Produtos", index=False)
        regions.to_excel(writer, sheet_name="Regiões", index=False)
        months.to_excel(writer, sheet_name="Mensal", index=False)

    wb = load_workbook(output_path)

    for name in wb.sheetnames:
        ws = wb[name]
        _style_header(ws)
        _autofit(ws)
        ws.freeze_panes = "A2"

    ws = wb["Dados Tratados"]
    headers = {cell.value: cell.column for cell in ws[1]}

    for name in ["Preço Unitário", "Receita Bruta", "Valor Desconto", "Receita Líquida"]:
        if name in headers:
            _currency_column(ws, headers[name])

    if "Desconto %" in headers:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=headers["Desconto %"]).number_format = "0.00%"

    if "Data" in headers:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=headers["Data"]).number_format = "dd/mm/yyyy"

    ws_summary = wb["Resumo"]
    for row in range(2, ws_summary.max_row + 1):
        indicator = ws_summary.cell(row=row, column=1).value
        if indicator in {"Receita Bruta", "Descontos Concedidos", "Receita Líquida", "Ticket Médio"}:
            ws_summary.cell(row=row, column=2).number_format = 'R$ #,##0.00'

    for name in ["Vendedores", "Produtos", "Regiões", "Mensal"]:
        ws_rank = wb[name]
        for cell in ws_rank[1]:
            if cell.value == "Receita_Liquida":
                _currency_column(ws_rank, cell.column)

    ws_sellers = wb["Vendedores"]
    if ws_sellers.max_row >= 2:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Receita Líquida por Vendedor"
        data = Reference(ws_sellers, min_col=4, min_row=1, max_row=ws_sellers.max_row)
        cats = Reference(ws_sellers, min_col=1, min_row=2, max_row=ws_sellers.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        ws_sellers.add_chart(chart, "F2")

    ws_month = wb["Mensal"]
    if ws_month.max_row >= 2:
        chart = LineChart()
        chart.title = "Evolução Mensal da Receita"
        data = Reference(ws_month, min_col=3, min_row=1, max_row=ws_month.max_row)
        cats = Reference(ws_month, min_col=1, min_row=2, max_row=ws_month.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 15
        ws_month.add_chart(chart, "E2")

    wb.save(output_path)
